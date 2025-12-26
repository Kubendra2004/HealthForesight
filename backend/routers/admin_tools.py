from fastapi import APIRouter, HTTPException, Depends, Response
from database.database import mongo_db, get_db
from auth.auth import get_current_user
from database.models_sql import User, UserRole
from sqlalchemy.orm import Session
import csv
import io
from typing import Optional

router = APIRouter(
    prefix="/admin",
    tags=["Admin Tools"]
)

# --- Data Export ---

@router.get("/export/patients")
async def export_patients(
    format: str = "csv",
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export all patients data as CSV or Excel"""
    if current_user.role != UserRole.admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Fetch patients from SQL DB
    from database.models_sql import User as SQLUser
    patients = db.query(SQLUser).filter(SQLUser.role == UserRole.patient).all()
    
    if format == "csv":
        # Generate CSV
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['ID', 'Username', 'Email', 'Active'])
        
        for patient in patients:
            writer.writerow([patient.id, patient.username, patient.email, patient.is_active])
        
        csv_content = output.getvalue()
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=patients.csv"}
        )
    
    elif format == "excel":
        # Generate Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Patients"
        
        # Header with styling
        headers = ['ID', 'Username', 'Email', 'Active']
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Add data
        for patient in patients:
            ws.append([patient.id, patient.username, patient.email, patient.is_active])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2
        
        # Save to bytes
        excel_output = io.BytesIO()
        wb.save(excel_output)
        excel_output.seek(0)
        
        return Response(
            content=excel_output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=patients.xlsx"}
        )
    
    return {"message": "Invalid format. Use 'csv' or 'excel'"}

@router.get("/export/appointments")
async def export_appointments(
    format: str = "csv",
    start_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Export appointments data as CSV or Excel"""
    if current_user.role != UserRole.admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Fetch from MongoDB
    appointments = await mongo_db.appointments.find().limit(1000).to_list(1000)
    
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Patient ID', 'Doctor ID', 'Date', 'Status', 'Reason', 'Type'])
        
        for appt in appointments:
            writer.writerow([
                appt.get('patient_id'),
                appt.get('doctor_id'),
                appt.get('date'),
                appt.get('status'),
                appt.get('reason'),
                appt.get('type', 'in-person')
            ])
        
        csv_content = output.getvalue()
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=appointments.csv"}
        )
    
    elif format == "excel":
        # Generate Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Appointments"
        
        # Header
        headers = ['Patient ID', 'Doctor ID', 'Date', 'Status', 'Reason', 'Type', 'Google Meet URL']
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for appt in appointments:
            ws.append([
                appt.get('patient_id'),
                appt.get('doctor_id'),
                str(appt.get('date')),
                appt.get('status'),
                appt.get('reason'),
                appt.get('type', 'in-person'),
                appt.get('meet_url', '')
            ])
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save
        excel_output = io.BytesIO()
        wb.save(excel_output)
        excel_output.seek(0)
        
        return Response(
            content=excel_output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=appointments.xlsx"}
        )
    
    return {"message": "Invalid format. Use 'csv' or 'excel'"}

# --- Report Builder (Simple) ---

@router.post("/reports/custom")
async def create_custom_report(
    report_name: str,
    collection: str,  # "patients", "appointments", etc.
    filters: dict,
    current_user: User = Depends(get_current_user)
):
    """
    Simple report builder.
    Example: {"status": "completed", "date": {"$gte": "2024-01-01"}}
    """
    if current_user.role != UserRole.admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        # Query MongoDB collection
        db_collection = getattr(mongo_db, collection)
        results = await db_collection.find(filters).limit(500).to_list(500)
        
        # Save report
        report = {
            "name": report_name,
            "collection": collection,
            "filters": filters,
            "created_by": current_user.username,
            "created_at": datetime.utcnow(),
            "result_count": len(results)
        }
        
        report_result = await mongo_db.custom_reports.insert_one(report)
        
        return {
            "id": str(report_result.inserted_id),
            "result_count": len(results),
            "message": "Report generated"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- Staff Rostering (Basic) ---

@router.post("/roster/generate")
async def generate_staff_roster(
    week_start: str,  # "2024-01-01"
    current_user: User = Depends(get_current_user)
):
    """
    Auto-generate weekly staff schedule based on predicted demand.
    Simplified implementation - in production would use optimization algorithm.
    """
    if current_user.role != UserRole.admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # TODO: Fetch predicted demand from Prophet models
    # TODO: Fetch staff availability
    # TODO: Run optimization to match staff to demand peaks
    
    # For now, return a simple structure
    roster = {
        "week_start": week_start,
        "shifts": [
            {"day": "Monday", "shift": "morning", "assigned_staff": [], "required_staff": 5},
            {"day": "Monday", "shift": "evening", "assigned_staff": [], "required_staff": 3},
            # ... more shifts
        ],
        "generated_at": datetime.utcnow()
    }
    
    result = await mongo_db.staff_rosters.insert_one(roster)
    return {"id": str(result.inserted_id), "message": "Roster generated"}

from datetime import datetime
